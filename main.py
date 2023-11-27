import openpyxl
from telethon.sync import TelegramClient
from telethon.tl.functions.channels import GetFullChannel


# Function to check if a Telegram channel is available and get the last post date
def check_telegram_channels(file_path, api_id, api_hash):
    # Load Excel file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Add a new column header for the results
    result_column_header = "Result"
    if result_column_header not in sheet[1]:
        sheet.cell(row=1, column=sheet.max_column + 1, value=result_column_header)

    # Initialize Telegram client
    client = TelegramClient('session_name', api_id, api_hash)
    client.connect()

    # Iterate through rows in the Excel file
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        channel_username = row[0]  # Assuming the channel username is in the first column

        try:
            # Get full information about the channel
            channel = client.get_entity(channel_username)
            full_channel = client(GetFullChannel(channel=channel))

            # Extract last post date from full_channel
            last_post_date = full_channel.full_chat.date

            result_message = f"Available - Last Post Date: {last_post_date}"
            sheet.cell(row=row_num, column=sheet.max_column, value=result_message)

            print(f"Channel: {channel_username} - {result_message}")
        except Exception as e:
            result_message = f"Unavailable - Error: {str(e)}"
            sheet.cell(row=row_num, column=sheet.max_column, value=result_message)

            print(f"Channel: {channel_username} - {result_message}")

    # Save changes to the Excel file
    wb.save(file_path)

    # Disconnect Telegram client
    client.disconnect()


# Set your Telegram API credentials
api_id = 'your_api_id'
api_hash = 'your_api_hash'

# Specify the path to your Excel file
excel_file_path = 'path/to/your/excel/file.xlsx'

# Call the function to check Telegram channels and save results
check_telegram_channels(excel_file_path, api_id, api_hash)
